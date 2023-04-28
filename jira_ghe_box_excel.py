''' a Python script to retrieve weekly issues status from JIRA and GHE boards, then append a new sheet in excel table'''
import logging
import re
import os
from io import BytesIO
from datetime import datetime, timedelta
import requests
import pytz
import openpyxl
from boxsdk import JWTAuth, Client
from jira import JIRA
from slack_sdk.webhook import WebhookClient

logger = logging.getLogger()
logging.basicConfig()
logger.setLevel(logging.INFO)

file_name = "Changes.xlsx"
slack_url_alert = os.getenv("SLACK_ALERT")
file_id = os.getenv("BOX_FILE_ID")
access_token_ghe = os.getenv("GHE_ACCESS_TOKEN")

headers = {"Authorization": f"Bearer {access_token_ghe}"}

def send_slack_message(text):
    """sending alert message to Slack"""
    webhook = WebhookClient(slack_url_alert)
    webhook.send(text=text)

def box_auth():
    """authentification with Box"""
    config = JWTAuth.from_settings_file('config.json')
    client = Client(config)
    return client

def download_table(client):
    """dowloading the excel table from Box"""
    try:
        file_content = client.file(file_id).content()
        wb = openpyxl.load_workbook(BytesIO(file_content))
        wb.save(file_name)
        with open(file_name, "wb") as f:
            f.write(file_content)
    except Exception:
        send_slack_message(text="ISSUES STATUS: Problem in download_table")

def create_sheet_if_not_exist():
    """Create a new sheet if it doesn't exist"""
    workbook = openpyxl.load_workbook(file_name)
    current_date = datetime.now().strftime("%Y-%m-%d")
    sheet_name = f"Status {current_date}"
    worksheet = None
    for sheet in workbook.sheetnames:
        if sheet == sheet_name:
            worksheet = workbook[sheet]
            break
    if worksheet is None:
        worksheet = workbook.create_sheet(sheet_name)
        worksheet.cell(row=1, column=1, value="Issue Key")
        worksheet.cell(row=1, column=2, value="Assignee")
        worksheet.cell(row=1, column=3, value="Summary")
        worksheet.cell(row=1, column=4, value="Status")
        worksheet.cell(row=1, column=5, value="Changes Since Last Week")
        worksheet.cell(row=1, column=6, value="RAG Status")
        worksheet.cell(row=1, column=7, value="Deadline")
    workbook.save(file_name)
    return sheet_name

def append_jira_data(sheet_name):
    """Append info from Jira"""
    password = os.getenv("USER_PASSWORD")
    username = os.getenv("USER_NAME")

    workbook = openpyxl.load_workbook(file_name)
    worksheet = workbook[sheet_name]
    # call JIRA
    options = {"server": "https://domain"}
    jira = JIRA(options, basic_auth=(username, password))
    jql=" your project Jira Query Language"
    issues = jira.search_issues(
        jql,
        maxResults=False,
        fields=["key", "assignee", "summary", "status", "comment", "labels", "duedate"],
    )
    for index, issue in enumerate(issues, start=2):
        has_update = False
        for comment in issue.fields.comment.comments:
            comment_date = datetime.strptime(comment.created, "%Y-%m-%dT%H:%M:%S.%f%z")
            utc_now = datetime.utcnow()
            utc_now_aware = utc_now.replace(tzinfo=pytz.utc)
            if (
                comment_date > utc_now_aware - timedelta(days=7)
                and "Changes since last week" in comment.body
            ):
                match = re.search(
                    r"Changes since last week:\s*(.*)", comment.body, re.DOTALL
                )
                if match:
                    result = match.group(1)
                    print(issue.fields.status)
                    worksheet.cell(
                        row=index,
                        column=1,
                        value=f"https://domain/browse/{issue.key}",
                    )
                    worksheet.cell(
                        row=index, column=2, value=issue.fields.assignee.displayName
                    )
                    worksheet.cell(row=index, column=3, value=issue.fields.summary)
                    worksheet.cell(row=index, column=4, value=issue.fields.status.name)
                    worksheet.cell(row=index, column=5, value=result)
                    if issue.fields.labels:
                        worksheet.cell(row=index, column=6, value=issue.fields.labels)
                    else:
                        worksheet.cell(row=index, column=6, value="No labels")
                    worksheet.cell(row=index, column=7, value=issue.fields.duedate)
                    has_update = True

        if has_update is False:
            worksheet.cell(
                row=index,
                column=1,
                value=f"https://domain/browse/{issue.key}",
            )
            worksheet.cell(row=index, column=2, value=issue.fields.assignee.displayName)
            worksheet.cell(row=index, column=3, value=issue.fields.summary)
            worksheet.cell(row=index, column=4, value=issue.fields.status.name)
            worksheet.cell(row=index, column=5, value="No report")
            for comment in issue.fields.comment.comments:
                if "RAG Status" in comment.body:
                    rag_status = re.search(r"RAG Status:(.*)", comment.body)
                    if rag_status:
                        worksheet.cell(row=index, column=6, value=rag_status.group(1))
                        break
                worksheet.cell(row=index, column=7, value=issue.fields.duedate)
    workbook.save(file_name)
    return sheet_name

def automation_issues_numbers_titles():
    """getting issue numbers and titles from GHE name folder board"""
    url = "https://domain/api/v3/repos/org/repo/issues"
    automation_issue_numbers = []
    automation_issue_titles = []
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        issues = response.json()
        for issue in issues:
            issue_name = issue["title"]
            automation_issue_titles.append(issue_name)
            issue_number = issue["number"]
            automation_issue_numbers.append(issue_number)
    return automation_issue_numbers, automation_issue_titles

def automation_label_ghe(automation_issue_numbers):
    """getting issues labels from GHE"""
    automation_list_labels = []
    headers_label = {
        "Accept": "application/vnd.github+json",
        "Authorization": f"Bearer {access_token_ghe}",
        "X-GitHub-Api-Version": "2022-11-28",
    }

    for number in automation_issue_numbers:
        url = f"https://domain/api/v3/repos/org/repo/issues/{number}/labels"
        response = requests.get(url, headers=headers_label)
        if response.status_code == 200:
            labels = response.json()
            label_names = [label["name"] for label in labels]
            label_names_str = ", ".join(label_names)
            if label_names_str == "":
                automation_list_labels.append("no label")
            else:
                automation_list_labels.append(label_names_str)
        else:
            print(f"Request failed with status code {response.status_code}.")
    return automation_list_labels

def automation_changes_last_week(automation_issue_numbers):
    """getting all the changes from the last comment GHE"""
    automation_list_comments = []
    for number in automation_issue_numbers:
        try:
            url = f"https://domain/api/v3/repos/org/repo/issues/{number}/comments"
            response = requests.get(url, headers=headers)
            comments = response.json()[-1]
            most_recent_comment = comments
            created_at = datetime.strptime(
                most_recent_comment["created_at"], "%Y-%m-%dT%H:%M:%SZ"
            )
            current_date = datetime.now()
            time_difference = current_date - created_at
            if time_difference <= timedelta(days=7):
                lines = most_recent_comment["body"].split("\n")
                result = "\n".join(lines[1:])
                automation_list_comments.append(result)
            else:
                automation_list_comments.append("No report")
        except:
            continue
    return automation_list_comments

def automation_get_issue_deadline(automation_issue_numbers):
    """getting the deadline from the initial GHE issue"""
    automation_list_deadlines = []
    for number in automation_issue_numbers:
        try:
            url = f"https://domain/api/v3/repos/org/repo/issues/{number}"
            response = requests.get(url, headers=headers)
            issue_data = response.json()
            issue_description = issue_data["body"]
            match = re.search(r"Deadline\s*(.*)$", issue_description, re.MULTILINE)
            if match:
                deadline = match.group(1).strip()
                result = deadline.split(": ")[-1]
                automation_list_deadlines.append(result)
            else:
                automation_list_deadlines.append("No deadline")
        except:
            continue
    return automation_list_deadlines

def automation_append_ghe_issues(
    automation_issue_numbers, automation_issue_titles, automation_list_labels, automation_list_comments, automation_list_deadlines
):
    """appending all the fields from GHE to final table"""
    workbook = openpyxl.load_workbook(file_name)
    last_sheet_name = workbook.sheetnames[-1]
    worksheet = workbook[last_sheet_name]
    next_row = worksheet.max_row + 1  # Get the next available row to write to
    for number, title, label, comment, deadline in zip(
        automation_issue_numbers, automation_issue_titles, automation_list_labels, automation_list_comments, automation_list_deadlines
    ):
        worksheet.cell(
            row=next_row,
            column=1,
            value=f"https://domain/org/repo/issues/{number}",
        )
        worksheet.cell(row=next_row, column=2, value="Galina Grudinina")
        worksheet.cell(row=next_row, column=3, value=title)  # summary
        worksheet.cell(row=next_row, column=4, value=label)  # status
        worksheet.cell(row=next_row, column=5, value=comment)  # comment
        worksheet.cell(row=next_row, column=6, value=label)
        worksheet.cell(row=next_row, column=7, value=deadline)
        next_row += 1
    workbook.save(file_name)

def report_automation_issues_numbers_titles():
    """getting issue numbers and titles from GHE name2 folder board"""
    url = "https://domain/api/v3/repos/org/repo2/issues"
    report_automation_issue_numbers = []
    report_automation_issue_titles = []
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        issues = response.json()
        for issue in issues:
            issue_name = issue["title"]
            report_automation_issue_titles.append(issue_name)
            issue_number = issue["number"]
            report_automation_issue_numbers.append(issue_number)
    return report_automation_issue_numbers, report_automation_issue_titles

def report_automation_label_ghe(report_automation_issue_numbers):
    """getting issues labels from GHE"""
    report_automation_list_labels = []
    headers_label = {
        "Accept": "application/vnd.github+json",
        "Authorization": f"Bearer {access_token_ghe}",
        "X-GitHub-Api-Version": "2022-11-28",
    }
    for number in report_automation_issue_numbers:
        url = f"https://domain/v3/repos/org/repo2/issues/{number}/labels"
        response = requests.get(url, headers=headers_label)
        if response.status_code == 200:
            labels = response.json()
            label_names = [label["name"] for label in labels]
            label_names_str = ", ".join(label_names)
            if label_names_str == "":
                report_automation_list_labels.append("no label")
            else:
                report_automation_list_labels.append(label_names_str)
        else:
            print(f"Request failed with status code {response.status_code}.")
    return report_automation_list_labels

def report_automation_changes_last_week(report_automation_issue_numbers):
    """getting all the changes from the last comment GHE"""
    report_automation_list_comments = []
    for number in report_automation_issue_numbers:
        try:
            url = f"https://domain/api/v3/repos/org/repo2/issues/{number}/comments"
            response = requests.get(url, headers=headers)
            comments = response.json()[-1]
            most_recent_comment = comments
            created_at = datetime.strptime(
                most_recent_comment["created_at"], "%Y-%m-%dT%H:%M:%SZ"
            )
            current_date = datetime.now()
            time_difference = current_date - created_at
            if time_difference <= timedelta(days=7):
                lines = most_recent_comment["body"].split("\n")
                result = "\n".join(lines[1:])
                report_automation_list_comments.append(result)
            else:
                report_automation_list_comments.append("No report")
        except:
            continue
    return report_automation_list_comments

def report_automation_get_issue_deadline(report_automation_issue_numbers):
    """getting the deadline from the initial GHE issue"""
    report_automation_list_deadlines = []
    for number in report_automation_issue_numbers:
        try:
            url = f"https://domain/api/v3/repos/org/repo2/issues/{number}"
            response = requests.get(url, headers=headers)
            issue_data = response.json()
            issue_description = issue_data["body"]
            match = re.search(r"Deadline\s*(.*)$", issue_description, re.MULTILINE)
            if match:
                deadline = match.group(1).strip()
                result = deadline.split(": ")[-1]
                report_automation_list_deadlines.append(result)
            else:
                report_automation_list_deadlines.append("No deadline")
        except:
            continue
    return report_automation_list_deadlines

def report_automation_append_ghe_issues(
    report_automation_issue_numbers, report_automation_issue_titles, report_automation_list_labels, report_automation_list_comments, report_automation_list_deadlines
):
    """appending all the fields from GHE to final table"""
    workbook = openpyxl.load_workbook(file_name)
    last_sheet_name = workbook.sheetnames[-1]
    worksheet = workbook[last_sheet_name]
    next_row = worksheet.max_row + 1  # Get the next available row to write to
    for number, title, label, comment, deadline in zip(
        report_automation_issue_numbers, report_automation_issue_titles, report_automation_list_labels, report_automation_list_comments, report_automation_list_deadlines
    ):
        worksheet.cell(
            row=next_row,
            column=1,
            value=f"https://domain/org/repo/issues/{number}",
        )
        worksheet.cell(row=next_row, column=2, value="Galina Grudinina")
        worksheet.cell(row=next_row, column=3, value=title)  # summary
        worksheet.cell(row=next_row, column=4, value=label)  # status
        worksheet.cell(row=next_row, column=5, value=comment)  # comment
        worksheet.cell(row=next_row, column=6, value=label)
        worksheet.cell(row=next_row, column=7, value=deadline)
        next_row += 1
    workbook.save(file_name)

def table_styling():
    """making the excel table look nicer aka wider and with a wrapped text"""
    workbook = openpyxl.load_workbook(file_name)
    last_sheet_name = workbook.sheetnames[-1]
    sheet = workbook[last_sheet_name]
    for column in sheet.columns:
        column_letter = column[0].column_letter
        column_width = 0
        for cell in column:
            cell_value = str(cell.value)
            max_line_length = max([len(line) for line in cell_value.split('\n')])
            line_count = cell_value.count("\n") + 1
            column_width = max(column_width, max_line_length)
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
            cell.font = openpyxl.styles.Font(name="Calibri", size=11)
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(border_style="thin", color="000000"),
                right=openpyxl.styles.Side(border_style="thin", color="000000"),
                top=openpyxl.styles.Side(border_style="thin", color="000000"),
                bottom=openpyxl.styles.Side(border_style="thin", color="000000"),
            )
            if column_letter in ['C', 'E']:
                sheet.column_dimensions[column_letter].width = 40
        if column_letter not in ['C', 'E']:
            adjusted_width = min(30, max(5, int(column_width * 1.2)))
            sheet.column_dimensions[column_letter].width = adjusted_width
        sheet.column_dimensions[column_letter].auto_size = True

    for row in sheet.rows:
        row_height = 0
        for cell in row:
            cell_value = str(cell.value)
            line_count = cell_value.count("\n") + 1
            row_height = max(row_height, 15 * line_count)
        row_number = row[0].row
        sheet.row_dimensions[row_number].height = row_height
    workbook.save(file_name)

def upload_to_box(client):
    """uploading a new sheet into table"""
    try:
        file_path = f"./{file_name}"
        client.file(file_id).update_contents(file_path)
    except Exception:
        send_slack_message(text="ISSUES STATUS: Problem in upload_to_box ")

if __name__ == "__main__":
    logger.info("Starting...")
    client = box_auth()
    download_table(client)
    sheet_name=create_sheet_if_not_exist()
    append_jira_data(sheet_name)
    automation_issue_numbers, automation_issue_titles=automation_issues_numbers_titles()
    automation_list_labels=automation_label_ghe(automation_issue_numbers)
    automation_list_comments=automation_changes_last_week(automation_issue_numbers)
    automation_list_deadlines=automation_get_issue_deadline(automation_issue_numbers)
    automation_append_ghe_issues(automation_issue_numbers, automation_issue_titles,automation_list_labels,automation_list_comments,automation_list_deadlines)
    report_automation_issue_numbers, report_automation_issue_titles=report_automation_issues_numbers_titles()
    report_automation_list_labels=report_automation_label_ghe(report_automation_issue_numbers)
    report_automation_list_comments=report_automation_changes_last_week(report_automation_issue_numbers)
    report_automation_list_deadlines=report_automation_get_issue_deadline(report_automation_issue_numbers)
    report_automation_append_ghe_issues(report_automation_issue_numbers, report_automation_issue_titles, report_automation_list_labels, report_automation_list_comments, report_automation_list_deadlines)
    table_styling()
    upload_to_box(client)
    logger.info("Completed")
