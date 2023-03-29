 """A simple script that retrieves changes from Jira board based on JQL and from GHE board, then appends a new sheet in existing excel table stored in Box""" 
import logging
import re
import os
import requests
import pytz
import openpyxl
from datetime import datetime, timedelta
from boxsdk import OAuth2, Client
from jira import JIRA
from slack_sdk.webhook import WebhookClient
from io import BytesIO

logger = logging.getLogger()
logging.basicConfig()
logger.setLevel(logging.INFO)

file_name = 'file_name.xlsx'
slack_url_alert=os.getenv("SLACK_ALERT")
client_id=os.getenv("BOX_CLIENT_ID")
client_secret=os.getenv("BOX_CLIENT_SECRET")
access_token=os.getenv("BOX_ACCESS_TOKEN")
file_id=os.getenv("BOX_FILE_ID")
access_token_ghe=os.getenv("GHE_ACCESS_TOKEN")
headers = {"Authorization": f"Bearer {access_token_ghe}"}

def send_slack_message(text):
    webhook = WebhookClient(slack_url_alert)
    webhook.send(text=text)

def box_auth():
    try:
        auth = OAuth2(
            client_id= client_id,
            client_secret=client_secret,
            access_token=access_token,
        )
        client = Client(auth)
        return client
    except Exception as e:
        send_slack_message(text="Problem in  box_auth")
        logger.error(e)

def download_table(client):
    try:
        file_content = client.file(file_id).content()
        wb = openpyxl.load_workbook(BytesIO(file_content))
        wb.save(file_name)
        with open(file_name,'wb') as f:
            f.write(file_content)
    except Exception as e:
        send_slack_message(text="Problem in download_table")
        logger.error(e)

def get_issues_numbers_titles():
    url='https://ghe.rakuten-it.com/api/v3/repos/org/repo/issues'
    issue_numbers=[]
    issue_titles=[]
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        issues = response.json()
        for issue in issues:
            issue_name = issue['title']
            issue_titles.append(issue_name)
            issue_number=issue['number']
            issue_numbers.append(issue_number)
    return issue_numbers, issue_titles

def create_sheet():
    try:
        password=os.getenv("USER_PASSWORD")
        username=os.getenv("USER_NAME")
        workbook = openpyxl.load_workbook(file_name)
        current_date = datetime.now().strftime('%Y-%m-%d')
        sheet_name = f'Status {current_date}'
        worksheet=None
        for sheet in workbook.sheetnames:
            if sheet == sheet_name:
                worksheet = workbook[sheet]
                break
        # Create a new sheet if it doesn't exist
        if worksheet is None:
            worksheet = workbook.create_sheet(sheet_name)
            worksheet.cell(row=1, column=1, value='Issue Key')
            worksheet.cell(row=1, column=2, value='Summary')
            worksheet.cell(row=1, column=3, value='Status')
        #call JIRA
        options = {'server': 'https://domain_name'}
        jira = JIRA(options, basic_auth=(username, password))
        jql="project=project name"
        issues = jira.search_issues(jql, maxResults=False, fields=["key","summary", "comment"])

        for index, issue in enumerate(issues, start=2):
            has_update = False
            for comment in issue.fields.comment.comments:
                comment_date = datetime.strptime(comment.created, '%Y-%m-%dT%H:%M:%S.%f%z')
                utc_now = datetime.utcnow()
                utc_now_aware = utc_now.replace(tzinfo=pytz.utc)
                if comment_date > utc_now_aware - timedelta(days=7) and 'Changes since last week' in comment.body:
                    match = re.search(r"Changes since last week:\s*(.*)", comment.body)
                    if match:
                        result = match.group(1)
                        worksheet.cell(row=index, column=1, value=issue.key)
                        worksheet.cell(row=index, column=2, value=issue.fields.summary)
                        worksheet.cell(row=index, column=3, value=result)
                        has_update = True

            if  has_update is False:
                worksheet.cell(row=index, column=1, value=issue.key)
                worksheet.cell(row=index, column=2, value=issue.fields.summary)
                worksheet.cell(row=index, column=3, value='No report')
        workbook.save(file_name)
    except Exception as e:
        send_slack_message(text="Problem in create_sheet")
        logger.error(e)

def append_table(issue_numbers, issue_titles):
    workbook = openpyxl.load_workbook(file_name)
    last_sheet_name = workbook.sheetnames[-1]
    worksheet = workbook[last_sheet_name]
    next_row = worksheet.max_row + 1  # Get the next available row to write to

    for number, title in zip(issue_numbers, issue_titles):
        worksheet.cell(row=next_row, column=1, value=number)
        worksheet.cell(row=next_row, column=2, value=title)
        url = f'https://domain_name/api/v3/repos/org/repo/issues/{number}/comments'
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            try:
                comment=[]
                last_comment = response.json()[-1]
                comment.append(last_comment['body'])
                print(last_comment['body'])
                for item in comment:
                    if "Changes for the" in item:
                        worksheet.cell(row=next_row, column=3, value=item)
                        next_row += 1  # Increment the next available row
                    else:
                        worksheet.cell(row=next_row, column=3, value="No report")
                        next_row += 1  # Increment the next available row
            except Exception as e:
                continue
    workbook.save(file_name)

def upload_to_box(client):
    try:
        file_path = f'./{file_name}'
        updated_file = client.file(file_id).update_contents(file_path)
    except Exception as e:
        send_slack_message(text="Problem in upload_to_box ")
        logger.error(e)

if __name__ == "__main__":
    logger.info("Starting...")
    client=box_auth()
    download_table(client)
    issue_numbers, issue_titles=get_issues_numbers_titles()
    create_sheet()
    append_table(issue_numbers, issue_titles)
    upload_to_box(client)
    logger.info("Completed")


